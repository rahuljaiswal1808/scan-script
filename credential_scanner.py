#!/usr/bin/env python3
"""
Comprehensive Credential Scanner
Detects hardcoded credentials in source code and exports findings to Excel.
"""

import os
import re
import json
import hashlib
import base64
import math
from pathlib import Path
from typing import List, Dict, Tuple, Set
from dataclasses import dataclass
from datetime import datetime
import argparse

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from tqdm import tqdm
except ImportError:
    print("Required packages not found. Install with:")
    print("pip install pandas openpyxl tqdm")
    exit(1)

@dataclass
class CredentialFinding:
    """Represents a detected credential finding"""
    file_path: str
    line_number: int
    line_content: str
    credential_type: str
    matched_pattern: str
    confidence_score: float
    severity: str
    context: str
    file_extension: str
    file_size: int = 0
    last_modified: str = ""
    file_permissions: str = ""
    relative_path: str = ""
    absolute_path: str = ""

class CredentialScanner:
    """Main credential scanner class"""
    
    def __init__(self, target_directory: str, output_file: str = "credential_findings.xlsx", config_path: str = "config.json"):
        self.target_directory = Path(target_directory)
        self.output_file = output_file
        self.findings: List[CredentialFinding] = []
        self.config = self._load_config(config_path)
        
        # Load settings from config or use defaults
        self.supported_extensions = set(self.config.get('scanning', {}).get('supported_extensions', [
            '.py', '.js', '.ts', '.jsx', '.tsx', '.java', '.c', '.cpp', '.h', '.hpp',
            '.cs', '.php', '.rb', '.go', '.rs', '.swift', '.kt', '.scala', '.sh',
            '.bash', '.zsh', '.fish', '.ps1', '.bat', '.cmd', '.yaml', '.yml',
            '.json', '.xml', '.properties', '.ini', '.cfg', '.conf', '.env',
            '.sql', '.r', '.m', '.pl', '.lua', '.dart', '.vue', '.html', '.css',
            '.scss', '.sass', '.less', '.toml', '.tf', '.dockerfile', '.md', '.txt'
        ]))
        
        self.excluded_files = set(self.config.get('scanning', {}).get('excluded_files', [
            'package-lock.json', 'yarn.lock', 'poetry.lock', 'Pipfile.lock',
            'composer.lock', 'Gemfile.lock', 'go.sum', 'Cargo.lock'
        ]))
        
        self.excluded_dirs = set(self.config.get('scanning', {}).get('excluded_directories', [
            '.git', '.svn', '.hg', 'node_modules', '__pycache__', '.pytest_cache',
            'venv', 'env', '.env', 'build', 'dist', 'target', 'bin', 'obj',
            '.vscode', '.idea', 'vendor', 'bower_components'
        ]))
        
        # Detection settings
        self.max_file_size = self.config.get('scanning', {}).get('max_file_size_mb', 10) * 1024 * 1024
        self.min_entropy = self.config.get('detection', {}).get('min_entropy', 4.5)
        self.min_length_for_entropy = self.config.get('detection', {}).get('min_length_for_entropy_check', 20)
        self.context_lines = self.config.get('detection', {}).get('context_lines', 2)
        
        # False positive reduction settings
        self.test_indicators = self.config.get('false_positive_reduction', {}).get('test_indicators', [
            'test', 'example', 'dummy', 'fake', 'mock', 'sample', 'placeholder'
        ])
        self.confidence_reduction_factor = self.config.get('false_positive_reduction', {}).get('confidence_reduction_factor', 0.5)
        
        # Whitelist settings
        whitelist_config = self.config.get('whitelist', {})
        self.whitelisted_files = set(whitelist_config.get('files', []))
        self.whitelisted_patterns = whitelist_config.get('patterns', [])
        self.whitelisted_file_patterns = whitelist_config.get('file_patterns', [])
        self.whitelisted_line_hashes = set(whitelist_config.get('line_hashes', []))
        
        self._init_patterns()
    
    def _load_config(self, config_path: str) -> Dict:
        """Load configuration from JSON file"""
        try:
            with open(config_path, 'r') as f:
                return json.load(f)
        except FileNotFoundError:
            print(f"Warning: Config file '{config_path}' not found. Using default settings.")
            return {}
        except json.JSONDecodeError as e:
            print(f"Warning: Invalid JSON in config file '{config_path}': {e}. Using default settings.")
            return {}
    
    def _init_patterns(self):
        """Initialize credential detection patterns"""
        self.patterns = {
            'AWS Access Key': [
                r'AKIA[0-9A-Z]{16}',
                r'aws_access_key_id\s*[:=]\s*["\']?AKIA[0-9A-Z]{16}["\']?',
                r'AWS_ACCESS_KEY_ID\s*[:=]\s*["\']?AKIA[0-9A-Z]{16}["\']?'
            ],
            'AWS Secret Key': [
                r'aws_secret_access_key\s*[:=]\s*["\']?[A-Za-z0-9/+=]{40}["\']?',
                r'AWS_SECRET_ACCESS_KEY\s*[:=]\s*["\']?[A-Za-z0-9/+=]{40}["\']?'
            ],
            'GitHub Token': [
                r'gh[pousr]_[A-Za-z0-9]{36}',
                r'github_token\s*[:=]\s*["\']?[A-Za-z0-9]{40}["\']?',
                r'GITHUB_TOKEN\s*[:=]\s*["\']?[A-Za-z0-9]{40}["\']?'
            ],
            'Google API Key': [
                r'AIza[0-9A-Za-z\-_]{35}',
                r'google_api_key\s*[:=]\s*["\']?AIza[0-9A-Za-z\-_]{35}["\']?'
            ],
            'Slack Token': [
                r'xox[baprs]-[0-9a-zA-Z\-]{10,48}',
                r'slack_token\s*[:=]\s*["\']?xox[baprs]-[0-9a-zA-Z\-]{10,48}["\']?'
            ],
            'JWT Token': [
                r'eyJ[A-Za-z0-9_-]*\.eyJ[A-Za-z0-9_-]*\.[A-Za-z0-9_-]*',
                r'jwt\s*[:=]\s*["\']?eyJ[A-Za-z0-9_-]*\.eyJ[A-Za-z0-9_-]*\.[A-Za-z0-9_-]*["\']?'
            ],
            'Database Password': [
                r'password\s*[:=]\s*["\'][^"\']{6,}["\']',
                r'passwd\s*[:=]\s*["\'][^"\']{6,}["\']',
                r'pwd\s*[:=]\s*["\'][^"\']{6,}["\']',
                r'db_password\s*[:=]\s*["\'][^"\']{6,}["\']',
                r'database_password\s*[:=]\s*["\'][^"\']{6,}["\']'
            ],
            'API Key (Generic)': [
                r'api_key\s*[:=]\s*["\'][A-Za-z0-9]{20,}["\']',
                r'apikey\s*[:=]\s*["\'][A-Za-z0-9]{20,}["\']',
                r'API_KEY\s*[:=]\s*["\'][A-Za-z0-9]{20,}["\']',
                r'APIKEY\s*[:=]\s*["\'][A-Za-z0-9]{20,}["\']'
            ],
            'Secret Key': [
                r'secret_key\s*[:=]\s*["\'][A-Za-z0-9/+=]{20,}["\']',
                r'SECRET_KEY\s*[:=]\s*["\'][A-Za-z0-9/+=]{20,}["\']',
                r'secret\s*[:=]\s*["\'][A-Za-z0-9/+=]{20,}["\']'
            ],
            'Private Key': [
                r'-----BEGIN\s+(RSA\s+)?PRIVATE\s+KEY-----',
                r'-----BEGIN\s+EC\s+PRIVATE\s+KEY-----',
                r'-----BEGIN\s+DSA\s+PRIVATE\s+KEY-----',
                r'-----BEGIN\s+OPENSSH\s+PRIVATE\s+KEY-----'
            ],
            'Certificate': [
                r'-----BEGIN\s+CERTIFICATE-----',
                r'-----BEGIN\s+PUBLIC\s+KEY-----'
            ],
            'Connection String': [
                r'mongodb://[^"\s]+',
                r'mysql://[^"\s]+',
                r'postgresql://[^"\s]+',
                r'redis://[^"\s]+',
                r'Server\s*=\s*[^;]+;\s*Database\s*=\s*[^;]+;\s*User\s+Id\s*=\s*[^;]+;\s*Password\s*=\s*[^;]+',
            ],
            'Email & Password': [
                r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}\s*[:=]\s*["\'][^"\']{6,}["\']'
            ],
            'Auth Header': [
                r'Authorization\s*:\s*Bearer\s+[A-Za-z0-9._-]+',
                r'Authorization\s*:\s*Basic\s+[A-Za-z0-9+/=]+',
                r'X-API-Key\s*:\s*[A-Za-z0-9._-]+'
            ],
            'Twilio': [
                r'AC[a-z0-9]{32}',
                r'SK[a-z0-9]{32}'
            ],
            'Stripe': [
                r'sk_live_[0-9a-zA-Z]{24}',
                r'sk_test_[0-9a-zA-Z]{24}',
                r'pk_live_[0-9a-zA-Z]{24}',
                r'pk_test_[0-9a-zA-Z]{24}'
            ],
            'SendGrid': [
                r'SG\.[a-zA-Z0-9_-]{22}\.[a-zA-Z0-9_-]{43}'
            ],
            'MailGun': [
                r'key-[0-9a-f]{32}'
            ],
            'Azure': [
                r'[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}'
            ]
        }
        
        # Add custom patterns from config
        custom_patterns = self.config.get('custom_patterns', {})
        self.patterns.update(custom_patterns)
        
        # Compile patterns for better performance
        self.compiled_patterns = {}
        for cred_type, patterns in self.patterns.items():
            self.compiled_patterns[cred_type] = [re.compile(pattern, re.IGNORECASE) for pattern in patterns]
    
    def calculate_entropy(self, string: str) -> float:
        """Calculate Shannon entropy of a string"""
        if not string:
            return 0
        
        entropy = 0
        for char in set(string):
            p = string.count(char) / len(string)
            entropy -= p * math.log2(p)
        return entropy
    
    def is_high_entropy_string(self, string: str, min_length: int = None, min_entropy: float = None) -> bool:
        """Check if string has high entropy (likely random/generated)"""
        if min_length is None:
            min_length = self.min_length_for_entropy
        if min_entropy is None:
            min_entropy = self.min_entropy
        
        if len(string) < min_length:
            return False
        return self.calculate_entropy(string) >= min_entropy
    
    def extract_context(self, lines: List[str], line_index: int, context_lines: int = None) -> str:
        """Extract context around the finding"""
        if context_lines is None:
            context_lines = self.context_lines
        
        start = max(0, line_index - context_lines)
        end = min(len(lines), line_index + context_lines + 1)
        context = []
        
        for i in range(start, end):
            prefix = ">>> " if i == line_index else "    "
            context.append(f"{prefix}{i+1}: {lines[i].rstrip()}")
        
        return "\n".join(context)
    
    def calculate_confidence_score(self, match_type: str, matched_text: str, context: str) -> Tuple[float, str]:
        """Calculate confidence score and severity for the finding"""
        base_score = 0.7
        severity = "Medium"
        
        # High confidence patterns
        high_confidence_types = {'AWS Access Key', 'AWS Secret Key', 'GitHub Token', 'JWT Token', 'Private Key'}
        if match_type in high_confidence_types:
            base_score = 0.9
            severity = "High"
        
        # Check for test/example indicators
        context_lower = context.lower()
        if any(indicator in context_lower for indicator in self.test_indicators):
            base_score *= self.confidence_reduction_factor
            severity = "Low"
        
        # Check for entropy
        if self.is_high_entropy_string(matched_text):
            base_score += 0.1
        
        # Check for common false positives
        false_positive_patterns = ['password', 'secret', 'key', 'token', 'api']
        if any(pattern in matched_text.lower() for pattern in false_positive_patterns):
            if len(matched_text) < 20:
                base_score *= 0.6
        
        # Adjust severity based on final score
        if base_score >= 0.8:
            severity = "High"
        elif base_score >= 0.6:
            severity = "Medium"
        else:
            severity = "Low"
        
        return min(base_score, 1.0), severity
    
    def is_whitelisted(self, file_path: Path, matched_text: str, line_content: str) -> bool:
        """Check if a finding should be ignored based on whitelist"""
        # Check if file is whitelisted
        relative_path = str(file_path.relative_to(self.target_directory))
        if relative_path in self.whitelisted_files:
            return True
        
        # Check file pattern matching
        for pattern in self.whitelisted_file_patterns:
            if re.match(pattern, relative_path):
                return True
        
        # Check if matched pattern is whitelisted
        for pattern in self.whitelisted_patterns:
            if re.search(pattern, matched_text, re.IGNORECASE):
                return True
        
        # Check line hash (for specific line content)
        line_hash = hashlib.sha256(line_content.encode()).hexdigest()
        if line_hash in self.whitelisted_line_hashes:
            return True
        
        return False
    
    def scan_file(self, file_path: Path) -> List[CredentialFinding]:
        """Scan a single file for credentials"""
        findings = []
        
        # Get file metadata
        try:
            stat_info = file_path.stat()
            file_size = stat_info.st_size
            last_modified = datetime.fromtimestamp(stat_info.st_mtime).isoformat()
            file_permissions = oct(stat_info.st_mode)[-3:]
        except Exception:
            file_size = 0
            last_modified = ""
            file_permissions = ""
        
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
            return findings
        
        for line_num, line in enumerate(lines, 1):
            line_content = line.rstrip()
            
            # Skip empty lines and comments in some languages
            if not line_content.strip() or line_content.strip().startswith(('#', '//', '/*', '*', '--')):
                continue
            
            for cred_type, patterns in self.compiled_patterns.items():
                for pattern in patterns:
                    matches = pattern.finditer(line_content)
                    for match in matches:
                        matched_text = match.group()
                        
                        # Check if this finding should be ignored
                        if self.is_whitelisted(file_path, matched_text, line_content):
                            continue
                        
                        context = self.extract_context(lines, line_num - 1)
                        
                        confidence, severity = self.calculate_confidence_score(
                            cred_type, matched_text, context
                        )
                        
                        relative_file_path = str(file_path.relative_to(self.target_directory))
                        absolute_file_path = str(file_path.resolve())
                        
                        finding = CredentialFinding(
                            file_path=absolute_file_path,  # Use full path as primary
                            line_number=line_num,
                            line_content=line_content,
                            credential_type=cred_type,
                            matched_pattern=matched_text,
                            confidence_score=confidence,
                            severity=severity,
                            context=context,
                            file_extension=file_path.suffix,
                            file_size=file_size,
                            last_modified=last_modified,
                            file_permissions=file_permissions,
                            relative_path=relative_file_path,
                            absolute_path=absolute_file_path
                        )
                        findings.append(finding)
        
        return findings
    
    def should_scan_file(self, file_path: Path) -> bool:
        """Determine if a file should be scanned"""
        # Check file extension
        if file_path.suffix.lower() not in self.supported_extensions:
            return False
        
        # Check if file is in excluded list
        if file_path.name in self.excluded_files:
            return False
        
        # Check file size (skip very large files)
        try:
            if file_path.stat().st_size > self.max_file_size:
                return False
        except:
            return False
        
        return True
    
    def scan_directory(self, show_progress: bool = True) -> None:
        """Recursively scan directory for credentials"""
        print(f"Starting credential scan of: {self.target_directory}")
        
        # First pass: collect all files to scan
        files_to_scan = []
        total_files = 0
        
        for root, dirs, files in os.walk(self.target_directory):
            # Remove excluded directories
            dirs[:] = [d for d in dirs if d not in self.excluded_dirs]
            
            root_path = Path(root)
            
            for file_name in files:
                file_path = root_path / file_name
                total_files += 1
                
                if self.should_scan_file(file_path):
                    files_to_scan.append(file_path)
        
        print(f"Found {total_files} total files, {len(files_to_scan)} will be scanned")
        
        # Second pass: scan files with progress bar
        if show_progress and len(files_to_scan) > 10:
            file_iterator = tqdm(files_to_scan, desc="Scanning files", unit="file")
        else:
            file_iterator = files_to_scan
        
        for file_path in file_iterator:
            if show_progress and len(files_to_scan) <= 10:
                print(f"Scanning: {file_path.relative_to(self.target_directory)}")
            
            file_findings = self.scan_file(file_path)
            self.findings.extend(file_findings)
            
            if file_findings and (not show_progress or len(files_to_scan) <= 10):
                print(f"  Found {len(file_findings)} potential credential(s)")
        
        print(f"\nScan complete!")
        print(f"Files scanned: {len(files_to_scan)}")
        print(f"Potential credentials found: {len(self.findings)}")
    
    def export_to_excel(self) -> None:
        """Export findings to Excel with formatting"""
        if not self.findings:
            print("No findings to export.")
            return
        
        # Convert findings to DataFrame
        data = []
        for finding in self.findings:
            data.append({
                'Full File Path': finding.file_path,
                'Relative Path': finding.relative_path,
                'Line Number': finding.line_number,
                'Credential Type': finding.credential_type,
                'Matched Pattern': finding.matched_pattern,
                'Confidence Score': finding.confidence_score,
                'Severity': finding.severity,
                'File Extension': finding.file_extension,
                'File Size (bytes)': finding.file_size,
                'Last Modified': finding.last_modified,
                'File Permissions': finding.file_permissions,
                'Line Content': finding.line_content,
                'Context': finding.context
            })
        
        df = pd.DataFrame(data)
        
        # Sort by severity and confidence
        severity_order = {'High': 3, 'Medium': 2, 'Low': 1}
        df['severity_rank'] = df['Severity'].map(severity_order)
        df = df.sort_values(['severity_rank', 'Confidence Score'], ascending=[False, False])
        df = df.drop('severity_rank', axis=1)
        
        # Create Excel file with formatting
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            # Summary sheet
            summary_data = {
                'Metric': ['Total Findings', 'High Severity', 'Medium Severity', 'Low Severity', 
                          'Average Confidence', 'Files Affected', 'Most Common Type'],
                'Value': [
                    len(self.findings),
                    len([f for f in self.findings if f.severity == 'High']),
                    len([f for f in self.findings if f.severity == 'Medium']),
                    len([f for f in self.findings if f.severity == 'Low']),
                    f"{sum(f.confidence_score for f in self.findings) / len(self.findings):.2f}",
                    len(set(f.file_path for f in self.findings)),
                    max(set(f.credential_type for f in self.findings), 
                        key=lambda x: len([f for f in self.findings if f.credential_type == x]))
                ]
            }
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Detailed findings
            df.to_excel(writer, sheet_name='Detailed Findings', index=False)
            
            # Format the sheets
            workbook = writer.book
            
            # Format summary sheet
            summary_sheet = workbook['Summary']
            summary_sheet.column_dimensions['A'].width = 20
            summary_sheet.column_dimensions['B'].width = 30
            
            # Format detailed findings sheet
            findings_sheet = workbook['Detailed Findings']
            findings_sheet.column_dimensions['A'].width = 60  # Full File Path
            findings_sheet.column_dimensions['B'].width = 30  # Relative Path
            findings_sheet.column_dimensions['D'].width = 20  # Credential Type
            findings_sheet.column_dimensions['E'].width = 30  # Matched Pattern
            findings_sheet.column_dimensions['M'].width = 50  # Line Content
            findings_sheet.column_dimensions['N'].width = 80  # Context
            
            # Apply conditional formatting for severity
            from openpyxl.styles import PatternFill
            
            red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            green_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
            
            for row in range(2, len(df) + 2):
                severity_cell = findings_sheet[f'G{row}']  # Severity is now column G
                if severity_cell.value == 'High':
                    for col in range(1, 15):  # Increased range for new columns
                        findings_sheet.cell(row=row, column=col).fill = red_fill
                elif severity_cell.value == 'Medium':
                    for col in range(1, 15):
                        findings_sheet.cell(row=row, column=col).fill = yellow_fill
                elif severity_cell.value == 'Low':
                    for col in range(1, 15):
                        findings_sheet.cell(row=row, column=col).fill = green_fill
        
        print(f"Results exported to: {self.output_file}")
    
    def export_to_json(self, output_file: str = None) -> None:
        """Export findings to JSON format for CI/CD integration"""
        if not self.findings:
            print("No findings to export.")
            return
        
        json_file = output_file or self.output_file.replace('.xlsx', '.json')
        
        # Convert findings to JSON-serializable format
        results = {
            "scan_info": {
                "timestamp": datetime.now().isoformat(),
                "target_directory": str(self.target_directory),
                "total_findings": len(self.findings),
                "files_scanned": len(set(f.file_path for f in self.findings))
            },
            "summary": {
                "high_severity": len([f for f in self.findings if f.severity == 'High']),
                "medium_severity": len([f for f in self.findings if f.severity == 'Medium']),
                "low_severity": len([f for f in self.findings if f.severity == 'Low']),
                "average_confidence": sum(f.confidence_score for f in self.findings) / len(self.findings),
                "most_common_type": max(set(f.credential_type for f in self.findings), 
                                     key=lambda x: len([f for f in self.findings if f.credential_type == x]))
            },
            "findings": []
        }
        
        for finding in self.findings:
            results["findings"].append({
                "full_file_path": finding.file_path,
                "relative_path": finding.relative_path,
                "absolute_path": finding.absolute_path,
                "line_number": finding.line_number,
                "line_content": finding.line_content,
                "credential_type": finding.credential_type,
                "matched_pattern": finding.matched_pattern,
                "confidence_score": finding.confidence_score,
                "severity": finding.severity,
                "file_extension": finding.file_extension,
                "file_size": finding.file_size,
                "last_modified": finding.last_modified,
                "file_permissions": finding.file_permissions,
                "context": finding.context
            })
        
        with open(json_file, 'w') as f:
            json.dump(results, f, indent=2)
        
        print(f"JSON results exported to: {json_file}")
    
    def export_to_csv(self, output_file: str = None) -> None:
        """Export findings to CSV format"""
        if not self.findings:
            print("No findings to export.")
            return
        
        csv_file = output_file or self.output_file.replace('.xlsx', '.csv')
        
        # Convert findings to DataFrame
        data = []
        for finding in self.findings:
            data.append({
                'Full File Path': finding.file_path,
                'Relative Path': finding.relative_path,
                'Line Number': finding.line_number,
                'Credential Type': finding.credential_type,
                'Matched Pattern': finding.matched_pattern,
                'Confidence Score': finding.confidence_score,
                'Severity': finding.severity,
                'File Extension': finding.file_extension,
                'File Size (bytes)': finding.file_size,
                'Last Modified': finding.last_modified,
                'File Permissions': finding.file_permissions,
                'Line Content': finding.line_content,
                'Context': finding.context.replace('\n', ' | ')  # Make single line for CSV
            })
        
        df = pd.DataFrame(data)
        df.to_csv(csv_file, index=False)
        print(f"CSV results exported to: {csv_file}")
    
    def export_to_html(self, output_file: str = None) -> None:
        """Export findings to HTML format with interactive features"""
        if not self.findings:
            print("No findings to export.")
            return
        
        html_file = output_file or self.output_file.replace('.xlsx', '.html')
        
        # Create HTML content
        html_content = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Credential Scanner Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .header {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 10px; margin-bottom: 20px; }}
        .summary {{ background: white; padding: 20px; border-radius: 10px; margin-bottom: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
        .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; }}
        .summary-item {{ text-align: center; padding: 15px; background: #f8f9fa; border-radius: 8px; }}
        .summary-item h3 {{ margin: 0; color: #333; }}
        .summary-item p {{ margin: 5px 0 0 0; font-size: 24px; font-weight: bold; }}
        .high {{ color: #dc3545; }}
        .medium {{ color: #fd7e14; }}
        .low {{ color: #28a745; }}
        table {{ width: 100%; border-collapse: collapse; background: white; border-radius: 10px; overflow: hidden; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background-color: #f8f9fa; font-weight: bold; }}
        tr:hover {{ background-color: #f5f5f5; }}
        .severity-high {{ background-color: #f8d7da; }}
        .severity-medium {{ background-color: #fff3cd; }}
        .severity-low {{ background-color: #d1f2eb; }}
        .code-context {{ font-family: monospace; background: #f8f9fa; padding: 10px; border-radius: 5px; white-space: pre-wrap; }}
        .file-path {{ font-weight: bold; color: #007bff; }}
        .credential-type {{ font-weight: bold; }}
        .confidence {{ font-weight: bold; }}
        .filter-container {{ background: white; padding: 15px; border-radius: 10px; margin-bottom: 20px; box-shadow: 0 2px 5px rgba(0,0,0,0.1); }}
        .filter-container input, .filter-container select {{ margin: 5px; padding: 8px; border: 1px solid #ddd; border-radius: 4px; }}
    </style>
    <script>
        function filterTable() {{
            const severityFilter = document.getElementById('severityFilter').value;
            const typeFilter = document.getElementById('typeFilter').value;
            const searchFilter = document.getElementById('searchFilter').value.toLowerCase();
            const rows = document.querySelectorAll('#findingsTable tbody tr');
            
            rows.forEach(row => {{
                const severity = row.cells[5].textContent;  // Severity is now column 5
                const type = row.cells[3].textContent;      // Type is now column 3
                const searchText = row.textContent.toLowerCase();
                
                const severityMatch = !severityFilter || severity === severityFilter;
                const typeMatch = !typeFilter || type === typeFilter;
                const searchMatch = !searchFilter || searchText.includes(searchFilter);
                
                row.style.display = (severityMatch && typeMatch && searchMatch) ? '' : 'none';
            }});
        }}
        
        function clearFilters() {{
            document.getElementById('severityFilter').value = '';
            document.getElementById('typeFilter').value = '';
            document.getElementById('searchFilter').value = '';
            filterTable();
        }}
    </script>
</head>
<body>
    <div class="header">
        <h1>🔍 Credential Scanner Report</h1>
        <p>Scan completed on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p>Target Directory: {self.target_directory}</p>
    </div>
    
    <div class="summary">
        <h2>Summary</h2>
        <div class="summary-grid">
            <div class="summary-item">
                <h3>Total Findings</h3>
                <p class="high">{len(self.findings)}</p>
            </div>
            <div class="summary-item">
                <h3>High Severity</h3>
                <p class="high">{len([f for f in self.findings if f.severity == 'High'])}</p>
            </div>
            <div class="summary-item">
                <h3>Medium Severity</h3>
                <p class="medium">{len([f for f in self.findings if f.severity == 'Medium'])}</p>
            </div>
            <div class="summary-item">
                <h3>Low Severity</h3>
                <p class="low">{len([f for f in self.findings if f.severity == 'Low'])}</p>
            </div>
            <div class="summary-item">
                <h3>Files Affected</h3>
                <p>{len(set(f.file_path for f in self.findings))}</p>
            </div>
            <div class="summary-item">
                <h3>Avg Confidence</h3>
                <p>{sum(f.confidence_score for f in self.findings) / len(self.findings):.2f}</p>
            </div>
        </div>
    </div>
    
    <div class="filter-container">
        <h3>Filters</h3>
        <input type="text" id="searchFilter" placeholder="Search..." onkeyup="filterTable()">
        <select id="severityFilter" onchange="filterTable()">
            <option value="">All Severities</option>
            <option value="High">High</option>
            <option value="Medium">Medium</option>
            <option value="Low">Low</option>
        </select>
        <select id="typeFilter" onchange="filterTable()">
            <option value="">All Types</option>"""
        
        # Add credential type options
        credential_types = sorted(set(f.credential_type for f in self.findings))
        for cred_type in credential_types:
            html_content += f'\n            <option value="{cred_type}">{cred_type}</option>'
        
        html_content += f"""
        </select>
        <button onclick="clearFilters()">Clear Filters</button>
    </div>
    
    <table id="findingsTable">
        <thead>
            <tr>
                <th>Full File Path</th>
                <th>Relative Path</th>
                <th>Line</th>
                <th>Type</th>
                <th>Matched Pattern</th>
                <th>Severity</th>
                <th>Confidence</th>
                <th>Context</th>
            </tr>
        </thead>
        <tbody>"""
        
        # Sort findings by severity and confidence
        sorted_findings = sorted(self.findings, 
                               key=lambda x: (
                                   {'High': 3, 'Medium': 2, 'Low': 1}[x.severity], 
                                   x.confidence_score
                               ), 
                               reverse=True)
        
        for finding in sorted_findings:
            severity_class = f"severity-{finding.severity.lower()}"
            html_content += f"""
            <tr class="{severity_class}">
                <td class="file-path" title="{finding.file_path}">{finding.file_path}</td>
                <td class="relative-path">{finding.relative_path}</td>
                <td>{finding.line_number}</td>
                <td class="credential-type">{finding.credential_type}</td>
                <td><code>{finding.matched_pattern}</code></td>
                <td class="confidence">{finding.severity}</td>
                <td>{finding.confidence_score:.2f}</td>
                <td><div class="code-context">{finding.context.replace('<', '&lt;').replace('>', '&gt;')}</div></td>
            </tr>"""
        
        html_content += """
        </tbody>
    </table>
</body>
</html>"""
        
        with open(html_file, 'w') as f:
            f.write(html_content)
        
        print(f"HTML results exported to: {html_file}")
    
    def generate_report(self) -> str:
        """Generate a text summary report"""
        if not self.findings:
            return "No credentials found."
        
        report = []
        report.append("=" * 60)
        report.append("CREDENTIAL SCAN REPORT")
        report.append("=" * 60)
        report.append(f"Scan Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append(f"Target Directory: {self.target_directory}")
        report.append(f"Total Findings: {len(self.findings)}")
        report.append("")
        
        # Severity breakdown
        high_count = len([f for f in self.findings if f.severity == 'High'])
        medium_count = len([f for f in self.findings if f.severity == 'Medium'])
        low_count = len([f for f in self.findings if f.severity == 'Low'])
        
        report.append("SEVERITY BREAKDOWN:")
        report.append(f"  High:   {high_count}")
        report.append(f"  Medium: {medium_count}")
        report.append(f"  Low:    {low_count}")
        report.append("")
        
        # Credential type breakdown
        type_counts = {}
        for finding in self.findings:
            type_counts[finding.credential_type] = type_counts.get(finding.credential_type, 0) + 1
        
        report.append("CREDENTIAL TYPES FOUND:")
        for cred_type, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            report.append(f"  {cred_type}: {count}")
        report.append("")
        
        # File breakdown
        file_counts = {}
        for finding in self.findings:
            file_counts[finding.file_path] = file_counts.get(finding.file_path, 0) + 1
        
        report.append("MOST AFFECTED FILES:")
        for file_path, count in sorted(file_counts.items(), key=lambda x: x[1], reverse=True)[:10]:
            report.append(f"  {file_path}: {count}")
        
        return "\n".join(report)

def main():
    parser = argparse.ArgumentParser(description='Scan source code for hardcoded credentials')
    parser.add_argument('directory', help='Directory to scan')
    parser.add_argument('-o', '--output', default='credential_findings.xlsx', 
                       help='Output file name (default: credential_findings.xlsx)')
    parser.add_argument('-c', '--config', default='config.json',
                       help='Configuration file path (default: config.json)')
    parser.add_argument('-f', '--format', choices=['excel', 'json', 'csv', 'html', 'all'], default='excel',
                       help='Output format: excel, json, csv, html, or all (default: excel)')
    parser.add_argument('-r', '--report', action='store_true', 
                       help='Print summary report to console')
    parser.add_argument('--exit-code', action='store_true',
                       help='Exit with non-zero code if high severity findings are found')
    parser.add_argument('--no-progress', action='store_true',
                       help='Disable progress bar display')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.directory):
        print(f"Error: Directory '{args.directory}' does not exist.")
        return 1
    
    # Create scanner and run scan
    scanner = CredentialScanner(args.directory, args.output, args.config)
    scanner.scan_directory(show_progress=not args.no_progress)
    
    if scanner.findings:
        # Export in requested format(s)
        if args.format in ['excel', 'all']:
            scanner.export_to_excel()
        if args.format in ['json', 'all']:
            scanner.export_to_json()
        if args.format in ['csv', 'all']:
            scanner.export_to_csv()
        if args.format in ['html', 'all']:
            scanner.export_to_html()
        
        if args.report:
            print("\n" + scanner.generate_report())
        
        # Exit with error code if high severity findings and --exit-code flag
        if args.exit_code:
            high_severity_count = len([f for f in scanner.findings if f.severity == 'High'])
            if high_severity_count > 0:
                print(f"\nExiting with code 1 due to {high_severity_count} high severity finding(s)")
                return 1
        
        return 0
    else:
        print("No credentials found in the scanned directory.")
        return 0

if __name__ == "__main__":
    sys.exit(main())